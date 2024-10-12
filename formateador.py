from Clases.productos import Producto
from Clases.categoria import Categoria
import os
from openpyxl import Workbook, load_workbook



class Formateador:

    # productos = Producto.obtener_productos()  # Este método obtiene todos los productos

    @staticmethod
    def obtener_ruta_escritorio():
        return os.path.join(os.path.expanduser("~"), "Desktop")

    @classmethod
    def crear_archivo_excel(cls):
        ruta_escritorio = cls.obtener_ruta_escritorio()
        archivo = os.path.join(ruta_escritorio, "IndiceMheing.xlsx")

        # Siempre creamos un nuevo archivo, sobrescribiendo el existente
        wb = Workbook()
        wb.save(archivo)
        print(f"Archivo 'IndiceMheing.xlsx' reescrito en el escritorio.")

        return archivo

    @classmethod
    def cargar_en_una_hoja(cls, archivo):
        wb = load_workbook(archivo)

        # Seleccionamos la hoja activa (por defecto) o creamos una si no existe
        if "Indice" not in wb.sheetnames:
            ws = wb.create_sheet(title="Indice")
        else:
            ws = wb["Indice"]

        # Encabezados generales
        ws.append(["Categoría", "Nombre del Producto", "Precio"])

        # Diccionario para agrupar productos por categoría
        productos = Producto.obtener_productos()
        productos_por_categoria = {}
        for producto in productos:
            # Suponiendo que cada producto tiene: nombre, link, precio, categoría, página
            nombre, link, precio, categoria, pagina = producto
            if categoria not in productos_por_categoria:
                productos_por_categoria[categoria] = []
            productos_por_categoria[categoria].append((nombre, precio))

        # Agregar categorías y sus productos a la hoja en un solo bloque
        categorias = Categoria.nombre_categorias()
        for categoria in categorias:
            productos = productos_por_categoria.get(categoria, [])

            # Agregar la categoría
            if productos:
                for producto, precio in productos:
                    ws.append([categoria, producto, precio])
                    categoria = ""  # Dejar en blanco la categoría después de la primera vez

        # Guardar el archivo Excel con los datos agregados
        wb.save(archivo)
        print("Datos cargados en una sola hoja del archivo.")


